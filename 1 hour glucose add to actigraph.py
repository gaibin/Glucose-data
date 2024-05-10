import pandas as pd
from openpyxl import load_workbook
import os

# Desktop path
desktop_path = os.path.join(os.path.expanduser('~Administrator'), 'Desktop')

# Read the first Excel file
file1_path = os.path.join(desktop_path, '1.xlsx')
df1 = pd.read_excel(file1_path)

# Read the second Excel file
file2_path = os.path.join(desktop_path, '2.xlsx')
df2 = pd.read_excel(file2_path, index_col=0)  # Use the first column as index

# Open the first Excel file to apply styles
wb1 = load_workbook(file1_path)
ws1 = wb1.active

# Iterate over each row in the first Excel sheet
for index, row in df1.iterrows():
    name = row['Subject']
    time = row['DateTime']
    
    try:
        # Find the column in the second Excel sheet where the matching blood sugar value is located and the next 12 consecutive values
        blood_sugar_column = df2[name]
        blood_sugar_values = blood_sugar_column[blood_sugar_column.index.get_loc(time):].tolist()[:13]
        
        # Get the index of the last empty cell column in the current row
        last_empty_column = len(df1.columns) + 1
        
        # Insert the 12 values into the last empty cell column
        for i, value in enumerate(blood_sugar_values):
            ws1.cell(row=index+2, column=last_empty_column + i, value=value)
    except KeyError:
        # Ignore unmatched names and times, continue processing the next match
        pass

# Save the modified first Excel file
output_file1_path = os.path.join(desktop_path, '1_with_blood_sugar.xlsx')
wb1.save(output_file1_path)

print("Blood sugar values have been filled in and saved as '1_with_blood_sugar.xlsx' on the desktop in the 'Administrator' folder.")
